"""
文档解析工具 - 支持 PDF, Word, Markdown, TXT, Excel
"""
from pathlib import Path
from typing import Optional
from logger import setup_logger

logger = setup_logger("tools.file_parser")


class FileParser:
    """解析各种格式的文档，提取文本内容"""
    
    @staticmethod
    async def parse(file_path: str, file_type: str = None) -> str:
        """解析文件，返回文本内容"""
        path = Path(file_path)
        if not file_type:
            file_type = path.suffix.lower()
        
        logger.info(f"解析文件: {path.name}, type={file_type}")
        
        parsers = {
            ".txt": FileParser._parse_txt,
            ".md": FileParser._parse_txt,
            ".pdf": FileParser._parse_pdf,
            ".doc": FileParser._parse_docx,
            ".docx": FileParser._parse_docx,
            ".xlsx": FileParser._parse_xlsx,
        }
        
        parser = parsers.get(file_type, FileParser._parse_txt)
        result = await parser(path)
        logger.info(f"文件解析完成: {path.name}, length={len(result)}")
        return result
    
    @staticmethod
    async def parse_bytes(content: bytes, filename: str) -> str:
        """从字节内容解析"""
        import tempfile, os
        suffix = Path(filename).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
            f.write(content)
            tmp_path = f.name
        try:
            return await FileParser.parse(tmp_path, suffix)
        finally:
            os.unlink(tmp_path)
    
    @staticmethod
    async def _parse_txt(path: Path) -> str:
        return path.read_text(encoding="utf-8")
    
    @staticmethod
    async def _parse_pdf(path: Path) -> str:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(path))
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            return text.strip()
        except Exception as e:
            return f"PDF解析失败: {e}"
    
    @staticmethod
    async def _parse_docx(path: Path) -> str:
        try:
            from docx import Document
            doc = Document(str(path))
            return "\n".join([p.text for p in doc.paragraphs])
        except Exception as e:
            return f"Word文档解析失败: {e}"
    
    @staticmethod
    async def _parse_xlsx(path: Path) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path))
            text_parts = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text_parts.append(" | ".join([str(cell) for cell in row if cell]))
            return "\n".join(text_parts)
        except Exception as e:
            return f"Excel解析失败: {e}"

file_parser = FileParser()
